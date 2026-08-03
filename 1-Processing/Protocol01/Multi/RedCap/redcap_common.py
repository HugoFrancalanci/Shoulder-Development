"""
Partie commune à tous les scripts d'extraction REDCap.
-----------------------------------------------------------------
Un script d'extraction par table (planning, preop, postop, ...) importe ce
module et ne définit que ce qui lui est propre : ses variables, la mise en
page de son en-tête, et comment construire une ligne à partir d'une ligne
REDCap.

Ce module gère ce qui ne change jamais d'une table à l'autre :
  - lecture du CSV REDCap
  - la liste des patients à extraire (ID_REDCAP), partagée par toutes les
    tables puisque c'est la même cohorte
  - résolution d'un id_redcap vers sa ligne (ou None si vide/introuvable)
  - conversion en nombre pour éviter le warning Excel "stocké en texte"
  - ligne vide avant chaque patient, ligne "R" si aucune donnée, ID affiché
    même quand la ligne est vide/introuvable ("ID Recap not available" si
    id_redcap n'a pas été renseigné)
"""

import datetime
import statistics

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------------
# CONFIGURATION PARTAGÉE
# -------------------------------------------------------------------------
INPUT_CSV = r"C:\Users\franc\OneDrive - Université de Genève\PhD Hugo\05_Ressources\02_Database\01_E02_Classification_rTSA\RedCap\Export\BASESDEDONNEESEPAULE_DATA_2026-07-08_1945.csv"

# ID REDCap des patients à extraire (même cohorte pour toutes les tables)
ID_REDCAP = [
    '20220803_104800_93',   # 1
    '20220613_094000_17',
    '20220601_112300_25',
    '20220919_113000_80',
    '20220905_114400_79',
    '20211129_100500_03',
    '20220815_123800_98',
    '20220817_122100_85',
    '20220406_083000_48',
    '20221111_082900_71',
    '20230208_084300_05',
    '20220831_104400_09',
    '20220729_103200_59',
    '20220902_081900_50',
    '20220914_083100_73',
    '20221109_084300_77',
    '20221007_084300_34',
    '20220930_105900_40',
    '20220930_135200_50',
    '20230111_112800_10',
    '20221104_083500_76',
    '20221121_131200_66',
    '20221107_094000_82',
    '20221123_105600_29',
    '20221123_082200_13',
    '',                     # 26 (vide)
    '20221130_105300_14',
    '20221128_151800_98',
    '20230111_085600_30',
    '20230113_084600_51',
    '20221214_084100_38',
    '20230329_090300_60',
    '20221130_083300_20',
    '20221205_093700_41',
    '20230109_093200_59',
    '20230125_082800_23',
    '20230118_082300_04',
    '20230116_094600_59',
    '20230123_093100_86',
    '20230227_124100_50',
    '20230125_140300_08',
    '20230213_135200_97',
    '20230320_094400_14',
    '20230322_083200_29',
    '20230310_111600_29',
    '20230403_092800_17',
    '20230510_085400_11',
    '20230522_112000_98',
    '20230421_110000_68',
    '20241227_091400_93',
    '20230522_141600_56',
    '20230612_100900_89',
    '20230517_082500_08',
    '20230619_121800_48',
    '20230612_130000_87',
    '20230619_154100_49',
    '20230614_083400_01',
    '20230920_090500_25',
    '20230731_100500_91',
    '20230626_100400_67',
    '20230918_093300_49',
    '20241125_121300_78',
    '20230804_083200_37',
    '20230901_085200_30',
    '20230920_114400_39',
    '20230823_132300_70',
    '20230925_112500_68',
    '20230911_113000_13',
    '20231002_095000_85',
    '20231009_093000_23',
    '20231002_122200_84',
    '20231030_093300_36',
    '20231129_083600_73',
    '20231127_100900_96',
    '20231129_111100_04',
    '20231211_104500_74',
    '20240108_114200_94',
    '20231120_115500_56',
    '20231120_092800_94',
    '20231204_094300_39',
    '',                     # 81 (vide)
    '20240115_092900_22',
    '',                     # 83 (vide)
    '20240117_082000_47',
    '20240110_085700_58',
    '20240212_093900_35',
    '20240122_093400_87',
    '20240115_121000_64',
    '20240108_092400_15',
    '20250108_123600_09',
    '20240205_093500_53',
    '20240306_084800_32',
    '20240228_084700_38',
    '20240214_115200_07',
    '20240419_090400_42',
    '20240408_093900_28',
    '20240325_092800_80',
    '20240327_114100_13',
    '20240320_085300_47',
    '20240419_113700_09',
    '20240327_083600_97',
    '20240318_101500_97',
    '20240410_112900_00',
    '20240506_093800_73',
    '20240429_092700_39',
    '20240605_090300_39',
    '20240513_093600_29',
    '20240429_121700_59',
    '20240603_103000_01',
    '20240719_084000_96',
    '20230809_160000_13',
    '20240527_112600_84',
    '20240612_084800_04',
    '20240617_123600_77',
    '20240617_094100_61',
    '20240619_084700_57',
    '20240621_083700_59',
    '20240729_115600_03',
    '20240909_093700_69',
    '20240916_094600_29',
    '20240805_093300_36',
    '20240715_112300_23',
    '20240729_093400_73',
    '20240918_091000_72',
    '20240916_123000_17',
    '20241106_084000_90',
    '20240923_093700_58',
    '20240819_093900_84',
    '20240918_115900_47',
    '20241009_113200_04',
    '20241028_093400_56',
    '20240902_094400_96',
    '20241009_085000_63',
    '20240930_094200_18',
    '',                     # 135 (vide)
    '20241014_093000_50',
    '20241007_094500_24',
    '20241016_104900_52',
    '20241111_094600_04',
    '20241120_085800_35',
    '20241104_092700_00',
    '20241113_114300_60',
    '20241120_113300_26',
    '20241104_114200_48',
    '20241125_093400_71',
    '20241113_090100_64',
    '20241127_120300_75',
    '20241216_134900_85',
    '20250120_094300_69',
    '20241202_130000_23',
    '20241211_084600_23',
    '20241202_104100_47',
    '20250115_084600_15',
    '20250108_101700_27',
    '20250113_104200_32',
    '20250106_094400_11',
    '20250604_131200_53',
    '20250305_112500_30',
    '20250210_094800_27',
    '20250212_084500_18',
    '20250210_122900_40',
    '20250303_093100_46',
    '20250303_130900_08',
    '20250324_140300_01',
    '20250324_094500_10',
    '20250409_083500_80',
    '20250414_100300_83',
    '20250514_084500_22',
    '20250416_132200_11',
    '20250430_103100_66',
    '20250416_103000_79',
    '20250428_143700_81',
    '20250428_115000_25',
    '20250528_084300_71',
    '20250512_093100_89',
    '20250519_093300_31',
    '20250602_120800_54',
    '20250523_084400_09',
    '20250604_085800_56',
    '20250521_161100_43',
    '20250528_110700_27',
    '20250602_094800_83',
    '20250620_133000_44',
]


# -------------------------------------------------------------------------
# LECTURE / RÉSOLUTION
# -------------------------------------------------------------------------
def load_csv(input_csv=INPUT_CSV):
    return pd.read_csv(input_csv, dtype=str)


def check_columns(df, fields):
    missing = [c for c in fields if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes introuvables dans le CSV : {missing}")


def get_patient_row(df, rid):
    """Lignes REDCap (DataFrame) pour un id_redcap, ou None si vide/absent.

    Un même id_redcap peut correspondre à PLUSIEURS lignes du CSV (events ou
    instruments répétés REDCap) : chaque ligne ne porte souvent qu'une
    partie des champs, le reste étant vide sur cette ligne-là. On renvoie
    donc toutes les lignes ; val() se charge d'aller chercher la bonne."""
    if rid == "":
        return None
    match = df[df["id_redcap"] == rid]
    if match.empty:
        print(f"ATTENTION - id_redcap introuvable dans le CSV : {rid}")
        return None
    return match


def val(rows, field):
    """Première valeur non vide de `field`, cherchée parmi toutes les
    lignes (events/instruments) d'un patient - le champ voulu peut être
    renseigné sur une ligne différente de celle des autres champs."""
    if rows is None:
        return ""
    for v in rows[field]:
        if not pd.isna(v) and str(v).strip() != "":
            return v
    return ""


def val_for_event(rows, field, event_name):
    """Comme val(), mais restreint aux lignes d'un événement REDCap donné
    (redcap_event_name). Utile pour les champs répétés à chaque visite
    (ex: cms_total, présent à la fois sur l'event "pre_operative_arm_1"
    et sur un event de suivi comme "fuy1_arm_1") où val() ne suffit pas
    puisqu'elle renverrait toujours la première valeur trouvée, peu
    importe l'événement."""
    if rows is None:
        return ""
    subset = rows[rows["redcap_event_name"] == event_name]
    for v in subset[field]:
        if not pd.isna(v) and str(v).strip() != "":
            return v
    return ""


def to_number(v):
    """Convertit en int/float si la valeur est numérique, sinon la laisse
    telle quelle (texte) - évite le warning Excel 'nombre stocké en texte'."""
    if not isinstance(v, str) or v.strip() == "":
        return v
    try:
        f = float(v.strip())
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


# -------------------------------------------------------------------------
# ÉCRITURE DU TABLEAU
# -------------------------------------------------------------------------
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def compute_reference_ranges(rows, n_reference=30, n_sd=2):
    """Calibre, colonne par colonne (index 0-based dans data_cols, PAS le
    tableau final avec ID), une plage [moyenne - n_sd*écart-type,
    moyenne + n_sd*écart-type] à partir des `n_reference` premiers patients
    AYANT DES DONNÉES (has_data=True), considérés fiables.

    rows : liste de (has_data, data_cols) dans l'ordre de ID_REDCAP.
    Une colonne n'est calibrée que si au moins 2 valeurs numériques sont
    trouvées parmi les patients de référence pour cette colonne (les
    colonnes texte/codes/labels ne sont jamais calibrées). Si l'écart-type
    est nul (valeur identique chez tous les patients de référence), cette
    valeur devient la seule acceptée (plage réduite à un point) plutôt que
    de ne rien calibrer.
    Retourne {col_idx: (low, high)}.
    """
    ref_rows = [dc for has_data, dc in rows if has_data][:n_reference]
    ranges = {}
    if not ref_rows:
        return ranges
    n_data_cols = len(ref_rows[0])
    for col_idx in range(n_data_cols):
        vals = [r[col_idx] for r in ref_rows if isinstance(r[col_idx], (int, float))]
        if len(vals) < 2:
            continue
        m, s = statistics.mean(vals), statistics.stdev(vals)
        ranges[col_idx] = (m, m) if s == 0 else (m - n_sd * s, m + n_sd * s)
    return ranges


def compute_reference_categories(rows, n_reference=30):
    """Équivalent catégoriel de compute_reference_ranges pour les colonnes
    texte (codes/labels traduits, marques, etc.) : calibre l'ensemble des
    valeurs distinctes vues parmi les n_reference premiers patients fiables
    (has_data=True). Une colonne où au moins une valeur numérique apparaît
    chez les patients de référence est ignorée ici (déjà couverte par
    compute_reference_ranges). "" et "R" (donnée manquante) ne comptent
    jamais comme une valeur - une donnée manquante n'est pas une anomalie.

    Limite : avec seulement n_reference patients, une catégorie légitime
    mais rare peut ne jamais apparaître dans le vocabulaire de référence et
    sera donc flaggée à tort - contrairement au cas numérique où une valeur
    proche de la moyenne reste "dans la plage" même si elle n'a jamais été
    vue exactement.

    Retourne {col_idx: set(valeurs connues)}.
    """
    ref_rows = [dc for has_data, dc in rows if has_data][:n_reference]
    categories = {}
    if not ref_rows:
        return categories
    n_data_cols = len(ref_rows[0])
    for col_idx in range(n_data_cols):
        if any(isinstance(r[col_idx], (int, float)) for r in ref_rows):
            continue  # colonne numérique, gérée par compute_reference_ranges
        vals = {r[col_idx] for r in ref_rows if isinstance(r[col_idx], str) and r[col_idx] not in ("", "R")}
        if vals:
            categories[col_idx] = vals
    return categories


def export_table(output_xlsx, sheet_title, header1, header2, merges, build_row,
                  id_redcap=ID_REDCAP, input_csv=INPUT_CSV, text_cols=None,
                  n_reference=30, n_sd=2):
    """
    header1, header2 : lignes d'en-tête (listes), même longueur.
    merges            : liste de (start_col, end_col) 1-based à fusionner sur header1.
    build_row(df, rid): doit retourner (has_data: bool, data_cols: list | None).
                        data_cols exclut la colonne ID (longueur = len(header2) - 1).
    text_cols         : liste de colonnes (1-based) à forcer en format Texte -
                        utile pour des codes/grades qui ressemblent à des
                        nombres ("0","1","2"...) mais n'en sont pas
                        (évite le warning Excel "nombre stocké en texte").
    n_reference/n_sd  : détection de valeurs hors plage/hors vocabulaire.
                        Calibré sur les n_reference premiers patients
                        fiables (has_data=True) : colonnes numériques ->
                        plage moyenne +/- n_sd écart-types (compute_reference_
                        ranges) ; colonnes texte/codes/labels -> ensemble des
                        valeurs déjà vues (compute_reference_categories).
                        Chaque cellule de TOUS les patients est ensuite
                        coloriée en vert (dans la plage/vocabulaire connu)
                        ou rouge (hors plage/jamais vue chez les patients de
                        référence). "R"/vide jamais colorié (donnée
                        manquante != anomalie).
    """
    df = load_csv(input_csv)

    n_cols = len(header2)
    blank_row = [""] * n_cols

    # Construit toutes les lignes en mémoire d'abord : la calibration des
    # plages/vocabulaires de référence a besoin des n_reference premiers
    # patients avant de savoir comment colorier n'importe quelle ligne (y
    # compris la 1re).
    built = [(rid, *build_row(df, rid)) for rid in id_redcap]
    ref_rows = [(hd, dc) for _, hd, dc in built]
    ranges = compute_reference_ranges(ref_rows, n_reference, n_sd)
    categories = compute_reference_categories(ref_rows, n_reference)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(header1)
    ws.append(header2)
    for start_col, end_col in merges:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    for rid, has_data, data_cols in built:
        ws.append(blank_row)

        if not has_data:
            id_display = rid if rid != "" else "ID RedCap not available"
            ws.append([id_display] + ["R"] * (n_cols - 1))
        else:
            ws.append([rid] + data_cols)
            row_idx = ws.max_row
            for col_idx, (low, high) in ranges.items():
                v = data_cols[col_idx]
                if not isinstance(v, (int, float)):
                    continue
                cell = ws.cell(row=row_idx, column=col_idx + 2)  # +1 colonne ID, +1 (1-based)
                cell.fill = GREEN_FILL if low <= v <= high else RED_FILL
            for col_idx, known_vals in categories.items():
                v = data_cols[col_idx]
                if not isinstance(v, str) or v in ("", "R"):
                    continue
                cell = ws.cell(row=row_idx, column=col_idx + 2)
                cell.fill = GREEN_FILL if v in known_vals else RED_FILL

        # Format Excel des dates (sinon Excel affiche "####" ou un format US)
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, datetime.date):
                cell.number_format = "DD.MM.YYYY"

    # Colonnes forcées en Texte (codes/grades ressemblant à des nombres)
    if text_cols:
        for col_idx in text_cols:
            for cell in ws[get_column_letter(col_idx)][2:]:  # sous les 2 lignes d'en-tête
                cell.number_format = "@"

    # Largeur de colonne auto (évite les "####" sur dates/nombres, colonnes
    # trop étroites pour le contenu)
    for col_idx in range(1, n_cols + 1):
        max_len = max(
            (len(str(cell.value)) for cell in ws[get_column_letter(col_idx)] if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 40))

    wb.save(output_xlsx)
    print(f"{len(id_redcap)} ligne(s) écrite(s) -> {output_xlsx}")
